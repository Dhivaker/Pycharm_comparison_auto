import os
import shutil
import xlrd
import xlwt
import matplotlib.pyplot as plt
import numpy as np
import itertools
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.lib.units import cm
from reportlab.lib import colors
import plotly.graph_objects as go
from reportlab.pdfgen import canvas
from datetime import datetime
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import openpyxl as openpyxl
from reportlab.lib.units import inch
import getpass
import PyPDF2
import subprocess
class PlotGen:
    #creating structure
    def __init__(self, plot_editor_excel,product_name,model_loc,datasheet_loc,model_type,hot_temp,model_level):
        #creating member variables
        self.dir_cur = os.getcwd()
        self.product_name = product_name
        self.hot_temp_value = hot_temp
        self.model_type = model_type
        self.datasheet_loc = datasheet_loc
        self.model_loc = model_loc
        self.model_level = model_level
        self.output_folder_name = ""
        #self.dir_output_path = os.path.join(self.dir_cur, self.output_folder_name)
        self.dir_netlist = os.path.join(self.dir_cur, "Netlist")
        self.dir_out_netlist = ""
        self.dir_out_plot = ""
        self.file_plot_editor = plot_editor_excel
        self.xlrd_plot_modifier = xlrd.open_workbook(plot_editor_excel)
        self.write_plot_modifier = openpyxl.load_workbook(plot_editor_excel)
        self.current_date = ""
        self.sim_writeSheet = self.write_plot_modifier["Sheet5"]
        self.excel_value = self.xlrd_plot_modifier.sheet_by_index(4)
        self.dict_simulated_files_data = {}
        self.dict_table_files_data = {}
        self.dict_plot_labels = {}
        self.dict_table_labels = {}
        self.dict_search_data ={}
        self.dict_data_sheet_value ={}
        self.dict_devices = {}
        self.load_plot_labels()
        self.load_table_labels()
        self.load_search_data()
        self.load_data_sheet_value()
        self.read_excel_sheet()
        #self.create_folders()



    def create_folders(self):
        output_folder = os.path.join(self.dir_cur,  "output{}_"+self.product_name)
        counter = 0
        output_folder.format(counter)
        while os.path.exists(output_folder.format(counter)):
           counter += 1
        output_folder = output_folder.format(counter)
        self.output_folder_name = output_folder
        self.dir_out_plot = os.path.join(output_folder, "output_plots")
        self.dir_out_netlist = os.path.join(output_folder, "netlist_files")
        # create dir for netlistfile
        os.makedirs(self.dir_out_netlist, mode=0o777,exist_ok=True)
        # create dir for plots
        os.makedirs(self.dir_out_plot, mode=0o777,exist_ok=True)



    def read_model_lib(self,product_name):
        for rows in range(1, self.sim_writeSheet.max_row):
            if product_name == self.excel_value.cell_value(rows, 1):
                with open(self.model_loc) as model_file:
                    model_file_lines = model_file.readlines()
                    header_index = 0
                    product_index = 0
                    for line_num in range(0, len(model_file_lines)):
                        if "Product" in model_file_lines[line_num] and "L2-model name" in model_file_lines[line_num] and "L3-model name" in model_file_lines[line_num]:
                            header_index = line_num
                            break
                    lines_to_read = header_index
                    while "End SIMPLIS Encryption" not in model_file_lines[lines_to_read]:
                        if product_name in model_file_lines[lines_to_read]:
                            product_index = lines_to_read
                            break
                        lines_to_read += 1
                    model_des = model_file_lines[product_index].split()
                    print(model_des)
                    product_name = model_des[1]
                    voltage_class = model_des[3]
                    current_class = model_des[4]
                    model_name = model_des[7]
                    model_name_column = self.sim_writeSheet.cell(rows + 1, 5)
                    model_name_column.value = model_name
                    voltage_class_column = self.sim_writeSheet.cell(rows + 1, 7)
                    voltage_class_column.value = int(voltage_class)
                    current_class_column = self.sim_writeSheet.cell(rows + 1, 6)
                    current_class_column.value = int(current_class)
                    library_path = self.model_loc
                    library_name = library_path.split("\\")
                    library = library_name[len(library_name) - 1]
                    library_column = self.sim_writeSheet.cell(rows + 1, 3)
                    library_column.value = library
                    print(product_name, voltage_class, current_class, model_name,library) # write back these printed values inside the excel
                    self.write_plot_modifier.save(self.file_plot_editor)

    def update_calibration_status(self):
        # for product_name in self.dict_devices.keys():
        #     if product_name == "SIGC100T65R3EA":      # SIGC100T65R3EA # IGC11T120X8L # IDC07D120X8L
        #
        self.create_folders()
        # self.output_folder_name = output_folder
        # self.dir_out_plot = os.path.join(output_folder, "output_plots")
        # self.dir_out_netlist = os.path.join(output_folder, "netlist_files")
        # self.read_model_lib(product_name)
        self.generate_net_files(self.product_name)
        self.load_simulation_data()
        self.find_x1_y1_value(self.product_name)
        self.modify_cies_data()

        # self.generate_plot()
        # self.load_table_data()
        # self._generate_table(self.product_name)
        # self._generate_pdf_report(self.product_name)
        excel_open = subprocess.Popen(["PDF_data_input.xlsx"], shell=True)
        excel_open.kill()

    def complete_report(self):


        # for product_name in self.dict_devices.keys():
        #     if product_name == "SIGC100T65R3EA":      # SIGC100T65R3EA # IGC11T120X8L # IDC07D120X8L
        self.create_folders()
        # self.output_folder_name = output_folder
        # self.dir_out_plot = os.path.join(output_folder, "output_plots")
        # self.dir_out_netlist = os.path.join(output_folder, "netlist_files")
        # self.read_model_lib(product_name)
        self.generate_net_files(self.product_name)
        self.load_simulation_data()
        self.find_x1_y1_value(self.product_name)
        self.modify_cies_data()
        self.generate_plot()
        self.load_table_data()
        self._generate_table(self.product_name)
        self._generate_pdf_report(self.product_name)


    def read_excel_sheet(self):
        for rows in range(3, self.excel_value.nrows):
            if self.product_name == self.excel_value.cell_value(rows,1):
                product_name = self.product_name
                self.dict_devices[product_name] = {}
                self.dict_devices[product_name]["library_name"] = self.excel_value.cell_value(rows, 2)
                self.dict_devices[product_name]["model_name"] = self.excel_value.cell_value(rows, 4)
                self.dict_devices[product_name]["current_class"] = self.excel_value.cell_value(rows,5)
                self.dict_devices[product_name]["voltage_class"] = self.excel_value.cell_value(rows, 6)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"] = {}
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["IC"] = self.excel_value.cell_value(rows,7)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["25°C"]={}
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["25°C"]["VCESat/VF_ref"] = self.excel_value.cell_value(rows,8)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["25°C"]["VCESat/VF_sim"] = self.excel_value.cell_value(rows,9)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["25°C"]["Error"] = self.excel_value.cell_value(rows,10)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["175°C"] = {}
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["175°C"]["VCESat/VF_ref"] = self.excel_value.cell_value(
                    rows, 11)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["175°C"]["VCESat/VF_sim"] = self.excel_value.cell_value(
                    rows, 12)
                self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["175°C"]["Error"] = self.excel_value.cell_value(rows, 13)
                self.dict_devices[product_name]["Transfer_characteristics"] = {}
                self.dict_devices[product_name]["Transfer_characteristics"]["IC"] = self.excel_value.cell_value(rows, 14)
                self.dict_devices[product_name]["Transfer_characteristics"]["175C"] = {}
                self.dict_devices[product_name]["Transfer_characteristics"]["175C"]["VGEth_ref"] = self.excel_value.cell_value(
                    rows, 15)
                self.dict_devices[product_name]["Transfer_characteristics"]["175C"]["VGEth_sim"] = self.excel_value.cell_value(
                    rows, 16)
                self.dict_devices[product_name]["Transfer_characteristics"]["175C"]["Error"] = self.excel_value.cell_value(rows, 17)
                self.dict_devices[product_name]["Capacitances"] = {}
                self.dict_devices[product_name]["Capacitances"]["VCE"] = self.excel_value.cell_value(rows, 18)
                self.dict_devices[product_name]["Capacitances"]["Cies"] ={}
                self.dict_devices[product_name]["Capacitances"]["Cies"]["cies_ref"] = self.excel_value.cell_value(
                    rows, 19)
                self.dict_devices[product_name]["Capacitances"]["Cies"]["cies_sim"] = self.excel_value.cell_value(
                    rows, 20)
                self.dict_devices[product_name]["Capacitances"]["Cies"]["Error"] = self.excel_value.cell_value(
                    rows, 21)
                self.dict_devices[product_name]["Capacitances"]["Crss"] = {}
                self.dict_devices[product_name]["Capacitances"]["Crss"]["crss_ref"] = self.excel_value.cell_value(
                    rows, 22)
                self.dict_devices[product_name]["Capacitances"]["Crss"]["crss_sim"] = self.excel_value.cell_value(
                    rows, 23)
                self.dict_devices[product_name]["Capacitances"]["Crss"]["Error"] = self.excel_value.cell_value(
                    rows, 24)
                self.dict_devices[product_name]["Capacitances"]["Coss"] ={}
                self.dict_devices[product_name]["Capacitances"]["Coss"]["coss_ref"] = self.excel_value.cell_value(
                    rows, 25)
                self.dict_devices[product_name]["Capacitances"]["Coss"]["coss_sim"] = self.excel_value.cell_value(
                    rows, 26)
                self.dict_devices[product_name]["Capacitances"]["Coss"]["Error"] = self.excel_value.cell_value(
                    rows, 27)
                # self.dict_devices[product_name]["location"] = self.excel_value.cell_value(
                #     rows, 27)
                # self.dict_devices[product_name]["Data_sheet_location"] = self.excel_value.cell_value(
                #     rows, 28)
                # self.dict_devices[product_name]["high_temp"] = self.excel_value.cell_value(
                #     rows, 29)
                return
        print("load the product value")





    def load_search_data(self):
        search_data = self.xlrd_plot_modifier.sheet_by_index(2)
        for rows in range(1,search_data.nrows):
            self.dict_search_data[search_data.cell_value(rows,0)] = search_data.cell_value(rows,1)



    def load_data_sheet_value(self):
        data_sheet_value = self.xlrd_plot_modifier.sheet_by_index(2)
        for rows in range(1,data_sheet_value.nrows):
            self.dict_data_sheet_value[data_sheet_value.cell_value(rows,0)] = data_sheet_value.cell_value(rows,2)

    def load_plot_labels(self):
        p_labels_sheet = self.xlrd_plot_modifier.sheet_by_index(1)
        #loading header
        headers = []
        for cell in p_labels_sheet.row(0):
            headers.append(cell.value)
        for rows in range(1,p_labels_sheet.nrows):
            self.dict_plot_labels[p_labels_sheet.cell_value(rows,0)]={}
            for cols in range(1,p_labels_sheet.ncols):
                self.dict_plot_labels[p_labels_sheet.cell_value(rows,0)][headers[cols]]=p_labels_sheet.cell_value(rows,cols)

    def load_table_labels(self):
        t_labels_sheet = self.xlrd_plot_modifier.sheet_by_index(3)
        #loading header
        headers = []
        for cell in t_labels_sheet.row(0):
            headers.append(cell.value)
        for rows in range(1,t_labels_sheet.nrows):
            self.dict_table_labels[t_labels_sheet.cell_value(rows,0)]={}
            for cols in range(1,t_labels_sheet.ncols):
                self.dict_table_labels[t_labels_sheet.cell_value(rows,0)][headers[cols]]=t_labels_sheet.cell_value(rows,cols)

    def get_user_input(self, product_name):
        # with open(self.file_netlist_modifier) as netlist_file:
        #     netlist_file_lines = netlist_file.readlines()
        user_input = {}
        if "IGC" in self.dict_devices[product_name]["model_name"]:
            user_input["location_IGBT"] = self.model_loc
            user_input["igbt_name"] = self.dict_devices[product_name]["model_name"]
        if "IDC" in self.dict_devices[product_name]["model_name"]:
            user_input["location_diode"] = self.model_loc
            user_input["diode_name"] = self.dict_devices[product_name]["model_name"]
        return user_input


    def __update_file_parameters(self, product_name):
        user_input = self.get_user_input(product_name)
        for files in os.listdir(self.dir_netlist):
            if files.endswith('.net') or files.endswith('.sxscr'):
                with open(os.path.join(self.dir_netlist, files)) as net_file:
                    net_file_lines = net_file.readlines()
                    for lines in range(0, len(net_file_lines)):
                        if "IGC" in self.dict_devices[product_name]["model_name"]:
                            if files.endswith('IGBT.sxscr') or files.startswith('design'):
                                if "<<LocationIGBT>>" in net_file_lines[lines]:
                                    net_file_lines[lines] = net_file_lines[lines].replace("<<LocationIGBT>>",
                                                                                          user_input["location_IGBT"])
                                if "<<IGBT_Name>>" in net_file_lines[lines]:
                                    net_file_lines[lines] = net_file_lines[lines].replace('<<IGBT_Name>>',
                                                                                          user_input["igbt_name"])
                        if "IDC" in self.dict_devices[product_name]["model_name"]:
                            if files.startswith('diode') or files.endswith('diode.sxscr'):
                                if "<<LocationDIODE>>" in net_file_lines[lines]:
                                    net_file_lines[lines] = net_file_lines[lines].replace("<<LocationDIODE>>",
                                                                                          user_input["location_diode"])

                                if "<<Diode_Name>>" in net_file_lines[lines]:
                                    net_file_lines[lines] = net_file_lines[lines].replace("<<Diode_Name>>",
                                                                                          user_input["diode_name"])
                        if "<<output>>" in net_file_lines[lines]:
                            output_folder = self.output_folder_name
                            folder_name = output_folder.split("/")
                            output = folder_name[len(folder_name) - 1]
                            net_file_lines[lines] = net_file_lines[lines].replace("<<output>>", output)
                        if "<<high_temp>>" in net_file_lines[lines]:
                            high_temp = str(self.hot_temp_value)
                            net_file_lines[lines] = net_file_lines[lines].replace("<<high_temp>>", high_temp)

                with open(os.path.join(self.dir_out_netlist, files), 'w') as output_file:
                    output_file.writelines(net_file_lines)

    def generate_net_files(self,product_name):
        self.__update_file_parameters(self.product_name)
        print('Started running simulations......Do some meditation')
        if "IGC" in self.dict_devices[product_name]["model_name"]:
            os.system("sim2 " + os.path.join(self.dir_out_netlist, "Script_all_simulations_IGBT.sxscr"))
        if "IDC" in self.dict_devices[product_name]["model_name"]:
            os.system("sim2 " + os.path.join(self.dir_out_netlist, "Script_all_simulations_diode.sxscr"))
        print('Fininshed running simulations')
        # print(os.path.join(netlist_dir,"Script_all_simulations.sxscr"))

    def find_closest_value(self,value,search_list):
        abs_diff = abs(search_list[0]-float(value))
        match_index = 0
        for itr in range(1,len(search_list)):
            if abs_diff > abs(float(value)-search_list[itr]):
                abs_diff = abs(float(value)-search_list[itr])
                match_index = itr
        return match_index


    def find_x1_y1_value(self,product_name):
        for plot_type in self.dict_simulated_files_data.keys():
            if plot_type == "output" or plot_type == "diode":
                for files in self.dict_simulated_files_data[plot_type].keys():
                    y1_value = float(self.dict_devices[product_name]["output_characteristics/Diode_Forward_characteristics"]["IC"])
                    self.dict_plot_labels[files]["y1value"] = y1_value
                    self.dict_plot_labels[files]["x1value"] = \
                        np.interp(y1_value, self.dict_simulated_files_data[plot_type][files]["y_axis"],
                                  self.dict_simulated_files_data[plot_type][files]["x_axis"])
                    x1_value = self.dict_plot_labels[files]["x1value"]
                    self.dict_table_labels[plot_type + ".png"]["2:1 value"] = y1_value
                    self.dict_table_labels[plot_type + ".png"]["2:2 value"] = y1_value
                    for rows in range(1, self.sim_writeSheet.max_row):
                        if product_name == self.excel_value.cell_value(rows,1):
                            today = datetime.now()
                            dt_string = today.strftime("%d/%m/%Y %H:%M:%S")
                            self.current_date = "%s" % dt_string
                            self.sim_writeSheet.cell(rows + 1, 4).value = self.current_date
                            if "output_25C.txt" in files or "diode_vf_25C.txt" in files:
                                self.dict_plot_labels[files]["legend"] = "TJ=25.0°C"
                                sim_data = self.sim_writeSheet.cell(rows+1,10)
                                sim_data.value = float("{:.2f}".format(x1_value))
                                data_sheet = self.excel_value.cell_value(rows, 8)
                                self.dict_table_labels[plot_type + ".png"]["3:1 value"] = data_sheet
                                self.dict_table_labels[plot_type+".png"]["4:1 value"] = sim_data.value
                                error_data = self.sim_writeSheet.cell(rows + 1, 11)
                                error_data.value = float("{:.2f}".format(float((sim_data.value-data_sheet)/sim_data.value)*100))
                                self.dict_table_labels[plot_type+".png"]["5:1 value"] = error_data.value
                            if "output_175C.txt" in files or "diode_vf_175C.txt" in files:
                                self.dict_plot_labels[files]["legend"] = "TJ="+str(self.hot_temp_value)+"°C"
                                sim_data = self.sim_writeSheet.cell(rows + 1, 13)
                                sim_data.value = float("{:.2f}".format(x1_value))
                                data_sheet = self.excel_value.cell_value(rows,11)
                                self.dict_table_labels[plot_type + ".png"]["3:2 value"] = data_sheet
                                self.dict_table_labels[plot_type+".png"]["4:2 value"] = sim_data.value
                                self.dict_table_labels[plot_type + ".png"]["1:2 value"] = "TJ="+str(self.hot_temp_value)+"°C"
                                error_data = self.sim_writeSheet.cell(rows+1,14)
                                error_data.value = float("{:.2f}".format(float((sim_data.value-data_sheet)/sim_data.value)*100))
                                self.dict_table_labels[plot_type+".png"]["5:2 value"] = error_data.value
                            self.write_plot_modifier.save(self.file_plot_editor)



                    # for rows in range(1, sim_writeSheet.max_row):
                    #     search_data = self.xlrd_plot_modifier.sheet_by_index(2)
                    #     filename = search_data.cell_value(rows,0)
                    #     if filename == files:
                    #         sim_data = sim_writeSheet.cell(rows+1,4)
                    #         sim_data.value = float("{:.2f}".format(x1_value))
                    #         self.write_plot_modifier.save(self.file_plot_editor)
                    # self.dict_plot_labels[files]["ty1"] = y1_value + 4
                    # self.dict_plot_labels[files]["tx1"] = x1_value - 5
            if plot_type == "data":
                for files in self.dict_simulated_files_data[plot_type].keys():
                    x1_value = float(self.dict_devices[product_name]["Capacitances"]["VCE"])
                    self.dict_plot_labels[files]["x1value"] = x1_value
                    self.dict_plot_labels[files]["y1value"] = \
                        np.interp(x1_value, self.dict_simulated_files_data[plot_type][files]["x_axis"],
                                  self.dict_simulated_files_data[plot_type][files]["y_axis"])
                    y1_value = self.dict_plot_labels[files]["y1value"]
                    self.dict_table_labels[plot_type + ".png"]["2:1 value"] = x1_value
                    self.dict_table_labels[plot_type + ".png"]["2:2 value"] = x1_value
                    self.dict_table_labels[plot_type + ".png"]["2:3 value"] = x1_value
                    for rows in range(1, self.sim_writeSheet.max_row):
                        if product_name == self.excel_value.cell_value(rows, 1):
                            if "data_cies.txt" in files:
                                sim_data = self.sim_writeSheet.cell(rows + 1, 21)
                                sim_data.value = float("{:.2e}".format(y1_value))
                                data_sheet = self.excel_value.cell_value(rows, 19)
                                self.dict_table_labels[plot_type + ".png"]["3:1 value"] = data_sheet
                                self.dict_table_labels[plot_type+".png"]["4:1 value"] = sim_data.value
                                self.write_plot_modifier.save(self.file_plot_editor)
                                if not data_sheet == "NA":
                                    error_data = self.sim_writeSheet.cell(rows + 1, 22)
                                    error_data.value = float("{:.2f}".format(
                                        float((sim_data.value-data_sheet)/sim_data.value) * 100))
                                    self.dict_table_labels[plot_type + ".png"]["5:1 value"] = error_data.value
                                else:
                                    self.sim_writeSheet.cell(rows + 1, 22).value = "NA"
                                    self.dict_table_labels[plot_type + ".png"]["5:1 value"] = "NA"
                            if "data_crss.txt" in files:
                                sim_data = self.sim_writeSheet.cell(rows + 1, 24)
                                sim_data.value = float("{:.2e}".format(y1_value))
                                data_sheet = self.excel_value.cell_value(rows, 22)
                                self.dict_table_labels[plot_type + ".png"]["3:2 value"] = data_sheet
                                self.dict_table_labels[plot_type+".png"]["4:2 value"] = sim_data.value
                                self.write_plot_modifier.save(self.file_plot_editor)
                                if not data_sheet == "NA":
                                    error_data = self.sim_writeSheet.cell(rows + 1, 25)
                                    error_data.value = float("{:.2f}".format(
                                        float((sim_data.value-data_sheet)/sim_data.value) * 100))
                                    self.dict_table_labels[plot_type + ".png"]["5:2 value"] = error_data.value
                                else:
                                    self.sim_writeSheet.cell(rows + 1, 25).value = "NA"
                                    self.dict_table_labels[plot_type + ".png"]["5:2 value"] = "NA"
                            if "data_coss.txt" in files:
                                sim_data = self.sim_writeSheet.cell(rows + 1, 27)
                                sim_data.value = float("{:.2e}".format(y1_value))
                                data_sheet = self.excel_value.cell_value(rows, 25)
                                self.dict_table_labels[plot_type + ".png"]["3:3 value"] = data_sheet
                                self.dict_table_labels[plot_type+".png"]["4:3 value"] = sim_data.value
                                self.write_plot_modifier.save(self.file_plot_editor)
                                if not data_sheet == "NA":
                                    error_data = self.sim_writeSheet.cell(rows + 1, 28)
                                    error_data.value = float("{:.2f}".format(
                                        float((sim_data.value-data_sheet)/sim_data.value) * 100))
                                    self.dict_table_labels[plot_type + ".png"]["5:3 value"] = error_data.value
                                else:
                                    self.sim_writeSheet.cell(rows + 1, 28).value = "NA"
                                    self.dict_table_labels[plot_type + ".png"]["5:3 value"] = "NA"
                            self.write_plot_modifier.save(self.file_plot_editor)

                    # self.dict_plot_labels[files]["ty1"] = y1_value + 4
                    # self.dict_plot_labels[files]["tx1"] = x1_value - 5
            if plot_type == "transfer":
                for files in self.dict_simulated_files_data[plot_type].keys():
                    if "transfer_25C.txt" in files:
                        y1_value = float(self.dict_devices[product_name]["Transfer_characteristics"]["IC"])
                        self.dict_plot_labels[files]["y1value"] = y1_value
                        self.dict_plot_labels[files]["x1value"] = \
                            np.interp(y1_value, self.dict_simulated_files_data[plot_type][files]["y_axis"],
                                      self.dict_simulated_files_data[plot_type][files]["x_axis"])
                        x1_value = self.dict_plot_labels[files]["x1value"]
                        self.dict_table_labels[plot_type + ".png"]["2:1 value"] = y1_value
                        self.dict_plot_labels[files]["legend"] = "TJ=25.0°C"
                        for rows in range(1, self.sim_writeSheet.max_row):
                            if product_name == self.excel_value.cell_value(rows, 1):
                                sim_data = self.sim_writeSheet.cell(rows + 1, 17)
                                sim_data.value = float("{:.2f}".format(x1_value))
                                data_sheet = self.excel_value.cell_value(rows, 15)
                                self.dict_table_labels[plot_type + ".png"]["4:1 value"] = sim_data.value
                                self.dict_table_labels[plot_type + ".png"]["3:1 value"] = data_sheet
                                error_data = self.sim_writeSheet.cell(rows + 1, 18)
                                error_data.value = float("{:.2f}".format(
                                    float((sim_data.value-data_sheet)/sim_data.value) * 100))
                                self.dict_table_labels[plot_type + ".png"]["5:1 value"] = error_data.value
                            self.write_plot_modifier.save(self.file_plot_editor)
                        # self.dict_plot_labels[files]["ty1"] = y1_value+4
                        # self.dict_plot_labels[files]["tx1"] = x1_value-5
                        if "transfer_175C.txt" in files:
                            self.dict_plot_labels[files]["legend"] = "TJ=" + str(
                                self.hot_temp_value) + "°C"


    # def find_x2_y2_value(self):
    #     for plot_type in self.dict_simulated_files_data.keys():
    #         if plot_type == "output" or plot_type == "diode":
    #             for files in self.dict_simulated_files_data[plot_type].keys():
    #                 x2_value = float(self.dict_data_sheet_value[files])
    #                 y2_value = float(self.dict_search_data[files])
    #                 self.dict_plot_labels[files]["y2value"] = y2_value
    #                 self.dict_plot_labels[files]["x2value"] = x2_value
    #                 # self.dict_plot_labels[files]["ty1"] = y1_value + 4
    #                 # self.dict_plot_labels[files]["tx1"] = x1_value - 5
    #         if plot_type == "data":
    #             for files in self.dict_simulated_files_data[plot_type].keys():
    #                 y2_value = float(self.dict_data_sheet_value[files])
    #                 x2_value = float(self.dict_search_data[files])
    #                 self.dict_plot_labels[files]["x2value"] = x2_value
    #                 self.dict_plot_labels[files]["y2value"] = y2_value
    #                 # self.dict_plot_labels[files]["ty1"] = y1_value + 4
    #                 # self.dict_plot_labels[files]["tx1"] = x1_value - 5
    #         if plot_type == "transfer":
    #             for files in self.dict_simulated_files_data[plot_type].keys():
    #                 if "transfer_25C.txt" in files:
    #                     x2_value = float(self.dict_data_sheet_value[files])
    #                     y2_value = float(self.dict_search_data[files])
    #                     self.dict_plot_labels[files]["y2value"] = y2_value
    #                     self.dict_plot_labels[files]["x2value"] = x2_value
    #                     # self.dict_plot_labels[files]["ty1"] = y1_value+4
    #                     # self.dict_plot_labels[files]["tx1"] = x1_value-5


    def modify_cies_data(self):
        for files in sorted(os.listdir(self.dir_out_netlist)):
            if "cies.txt" in files:
                self.dict_simulated_files_data["data"]["data_cies.txt"]["x_axis"] = self.dict_simulated_files_data["data"]["data_coss.txt"]["x_axis"]
                y_axis_last_index = len(self.dict_simulated_files_data["data"]["data_cies.txt"]["y_axis"]) - 1
                y_axis_last_value = self.dict_simulated_files_data["data"]["data_cies.txt"]["y_axis"][y_axis_last_index]
                self.dict_simulated_files_data["data"]["data_cies.txt"]["y_axis"] = list(itertools.repeat(y_axis_last_value, len(self.dict_simulated_files_data["data"]["data_cies.txt"]["x_axis"])))

    def load_simulation_data(self):
        for files in sorted(os.listdir(self.dir_out_netlist)):
            if files.endswith('.txt'):
                print("Loading file:", files)
                file_index = files.split('_')[0]
                if file_index not in self.dict_simulated_files_data:
                    self.dict_simulated_files_data[file_index] = {}
                self.dict_simulated_files_data[file_index][files]={}
                self.dict_simulated_files_data[file_index][files]["x_axis"] = []
                self.dict_simulated_files_data[file_index][files]["y_axis"] = []

                with open(os.path.join(self.dir_out_netlist, files)) as txt_file:
                    raw_data = txt_file.readlines()
                    if (len(raw_data[0].split()) > 2):
                        index_x = 1; index_y = 2
                    else:
                        index_x = 0; index_y = 1

                    for itr in range(1, len(raw_data)):
                        self.dict_simulated_files_data[file_index][files]["x_axis"].append(float(raw_data[itr].split()[index_x].strip()))
                        self.dict_simulated_files_data[file_index][files]["y_axis"].append(float(raw_data[itr].split()[index_y].strip()))



    def load_table_data(self):
        for images in sorted(os.listdir(self.dir_out_plot)):
            if images.endswith('.png') and not images.startswith('table') and not images.startswith('model'):
                print("Loading images:", images)
                self.dict_table_files_data[images] = {}


    def get_plabels(self,filename):
        return self.dict_plot_labels[filename]

    def get_tlabels(self,imagename):
        return self.dict_table_labels[imagename]


    def generate_plot(self):
        for plot_type in self.dict_simulated_files_data.keys():
            for files in self.dict_simulated_files_data[plot_type].keys():
                labels = self.get_plabels(files)
                plt.plot(self.dict_simulated_files_data[plot_type][files]["x_axis"],
                         self.dict_simulated_files_data[plot_type][files]["y_axis"],
                         label=labels["legend"], color=labels["color"], linewidth=labels["width"])
                plt.grid(True)
                plt.ylabel(labels["y_label"],fontsize= 14)
                plt.xlabel(labels["x_label"],fontsize= 14)
                plt.yscale(labels["scale"])
                #plt.title(labels["title"])
                plt.legend()
                if "output" == plot_type or "diode" == plot_type:
                    plt.axhline(labels["y1value"], color="k", linestyle="--")
                    plt.axvline(labels["x1value"], color="k", linestyle="--")
                    plt.xlim(left=0)
                    plt.ylim(bottom=0)
                    # plt.annotate(labels["text"], xytext=(labels["tx1"], labels["ty1"]),
                    #              xy=(labels["x1value"], labels["y1value"]),
                    #              arrowprops=dict(facecolor='black', shrink=0.05))
                if "transfer_25C" in files:
                    plt.axhline(labels["y1value"], color="k", linestyle="--")
                    plt.axvline(labels["x1value"], color="k", linestyle="--")
                    plt.xlim(left=0)
                    plt.ylim(bottom=0)
                    # plt.annotate(labels["text"], xytext=(labels["tx1"], labels["ty1"]),
                    #              xy=(labels["x1value"], labels["y1value"]),
                    #              arrowprops=dict(facecolor='black', shrink=0.05))
                if "data" in files:
                    # plt.yticks(np.arange(10E-14, 10E-9, 10E-1/2))
                    plt.axhline(labels["y1value"], color="k", linestyle="--")
                    plt.axvline(labels["x1value"], color="k", linestyle="--")
                    plt.xlim(left=0)
                    plt.ylim(10E-13, 10E-8)

                # plt.ylim(labels["ymin"], labels["ymax"])

            image_file = plot_type + ".png"
            plt.savefig(os.path.join(self.dir_out_plot, image_file), dpi= 300)
            plt.clf()



    def _generate_table(self,product_name):

        product = self.product_name
        model = self.dict_devices[product_name]["model_name"]
        library = self.dict_devices[product_name]["library_name"]
        voltage = self.dict_devices[product_name]["voltage_class"]
        current = self.dict_devices[product_name]["current_class"]
        fig = go.Figure(data=[go.Table(
            columnwidth=[250, 400],
            cells=dict(values=[
                ['Product name', 'Model name', 'Model library', 'Voltage class(V)', 'Current class(A)', 'Model level', 'Model type'],
                [product, model, library, voltage, current, self.model_level, self.model_type]], # get model level and type as input from user
                line_color='black',
                fill_color='white',
                align='center',
                font=dict(color='black', size=22),
                height=40))
        ])

        fig.update_layout(width=1000, height=800)
        fig.write_image(os.path.join(self.dir_out_plot, "model_det.png"))

        for images in self.dict_table_files_data.keys():
            labels = self.get_tlabels(images)
            if "output.png" == images or "diode.png" == images:
                fig = go.Figure(data=[go.Table(
                    columnwidth=[250, 150, 350, 350, 200],
                    header=dict(values=[labels["t1st column"], labels["t2nd column"], labels["t3rd column"],
                                        labels["t4th column"], labels["t5th column"]],
                                line_color='black',
                                fill_color='white',
                                align='center',
                                font=dict(color='black', size=22),
                                height=60),
                    cells=dict(values=[[labels["1:1 value"], labels["1:2 value"]],
                                       [labels["2:1 value"], labels["2:2 value"]],
                                       [labels["3:1 value"], labels["3:2 value"]],
                                       [labels["4:1 value"], labels["4:2 value"]],
                                       [labels["5:1 value"], labels["5:2 value"]]],
                               line_color='black',
                               fill= dict(color=['white',
                                                 'White',
                                                 'White',
                                                 'White',
                                                 ['rgba(0,250,0,0.8)' if not type(labels["5:1 value"])==str and labels["5:1 value"] >=-10 and labels["5:1 value"] <=10 else
                                                 'white' if type(labels["5:1 value"])== str else 'rgba(250, 0, 0, 0.8)',

                                                 'rgba(0,250,0,0.8)' if not type(labels["5:2 value"])==str and labels["5:2 value"] >=-10 and labels["5:2 value"] <=10 else
                                                 'white' if type(labels["5:2 value"])== str else 'rgba(250, 0, 0, 0.8)']]
                                                    ),
                               align='center',
                               font=dict(color='black', size=22),
                               height=40))
                ])

                fig.update_layout(width=1000, height=800)
                table_file = "table_" + images
                fig.write_image(os.path.join(self.dir_out_plot, table_file))

            if "transfer.png" in images:
                fig = go.Figure(data=[go.Table(
                    columnwidth=[250, 150, 350, 350, 200],
                    header=dict(values=[labels["t1st column"], labels["t2nd column"], labels["t3rd column"],
                                        labels["t4th column"], labels["t5th column"]],
                                line_color='black',
                                fill_color='white',
                                align='center',
                                font=dict(color='black', size=22),
                                height=60),
                    cells=dict(values=[[labels["1:1 value"]],
                                       [labels["2:1 value"]],
                                       [labels["3:1 value"]],
                                       [labels["4:1 value"]],
                                       [labels["5:1 value"]]],
                               line_color='black',
                               fill=dict(color=['white',
                                                'White',
                                                'White',
                                                'White',
                                                ['rgba(0,250,0,0.8)' if not type(labels["5:1 value"])==str and labels["5:1 value"] >=-10 and labels["5:1 value"] <=10 else
                                                 'white' if type(labels["5:1 value"])== str else 'rgba(250, 0, 0, 0.8)']]
                                         ),
                               align='center',
                               font=dict(color='black', size=22),
                               height=40))
                ])

                fig.update_layout(width=1000, height=800)
                fig.write_image(os.path.join(self.dir_out_plot, "table_transfer.png"))

            if "data.png" in images:
                fig = go.Figure(data=[go.Table(
                    columnwidth=[250, 150, 350, 350, 200],
                    header=dict(values=[labels["t1st column"], labels["t2nd column"], labels["t3rd column"],
                                        labels["t4th column"], labels["t5th column"]],
                                line_color='black',
                                fill_color='white',
                                align='center',
                                font=dict(color='black', size=22),
                                height=60),
                    cells=dict(values=[[labels["1:1 value"], labels["1:2 value"], labels["1:3 value"]],
                                       [labels["2:1 value"], labels["2:2 value"], labels["2:3 value"]],
                                       [labels["3:1 value"], labels["3:2 value"], labels["3:3 value"]],
                                       [labels["4:1 value"],
                                        labels["4:2 value"],
                                        labels["4:3 value"]],
                                       [labels["5:1 value"],
                                        labels["5:2 value"],
                                        labels["5:3 value"]]],
                               line_color='black',
                               fill=dict(color=['white',
                                                'White',
                                                'White',
                                                'White',
                                                ['rgba(0,250,0,0.8)' if not type(labels["5:1 value"])==str and labels["5:1 value"] >=-10 and labels["5:1 value"] <=10 else
                                                 'white' if type(labels["5:1 value"])== str else 'rgba(250, 0, 0, 0.8)',

                                                 'rgba(0,250,0,0.8)' if not type(labels["5:2 value"])==str and labels["5:2 value"] >=-10 and labels["5:2 value"] <=10 else
                                                 'white' if type(labels["5:2 value"])== str else 'rgba(250, 0, 0, 0.8)',

                                                 'rgba(0,250,0,0.8)' if not type(labels["5:3 value"])==str and labels["5:3 value"] >=-10 and labels["5:3 value"] <=10 else
                                                 'white' if type(labels["5:3 value"])== str else 'rgba(250, 0, 0, 0.8)']]
                                                ),
                               align='center',
                               font=dict(color='black', size=22),
                               height=40))
                ])
                fig.update_layout(width=1000, height=800)
                fig.write_image(os.path.join(self.dir_out_plot, "table_data.png"))




    def _generate_pdf_report(self,product_name):

        pdf_filename = self.product_name+'_Baredie_report{}.pdf'
        counter = 0
        while os.path.isfile(pdf_filename.format(counter)):
            counter += 1
        pdf_filename = pdf_filename.format(counter)
        pdf_folder = os.path.join(self.dir_out_plot, pdf_filename)
        pdf = canvas.Canvas(pdf_folder)
        logo = 'infineon_logo.png'
        front_page = 'frontpage.png'
        array_sort = []
        if "IGC" in self.dict_devices[product_name]["model_name"]:
            array_sort = ["transfer.png","output.png","data.png"]
        if "IDC" in self.dict_devices[product_name]["model_name"]:
            array_sort = ["diode.png"]
        for images in array_sort:
                if "transfer.png" in images or "diode.png" in images:
                    pdf.drawImage(front_page, 0, -3,width=21.15*cm, height=30.08*cm )
                    #pdf.drawImage(logo, 450, 780, width=80, height=40, mask='auto')
                    pdfmetrics.registerFont(
                        TTFont('source', 'SourceSansPro-Regular.ttf')
                    )
                    pdfmetrics.registerFont(
                        TTFont('source-bold', 'SourceSansPro-Bold.ttf')
                    )
                    pdf.setFont("source", 28)
                    pdf.drawCentredString(290, 740, "Compact model")
                    pdf.drawString(185, 705, "calibration report")
                    pdf.setFont("source", 24)

                    pdf.drawCentredString(290, 520, "Device: "+self.product_name)
                    #pdf.drawCentredString(270, 560, "Diode: IDC07D120X8L_L2")
                    pdf.setFont("source", 14)
                    today = datetime.now()
                    dt_string = today.strftime("%d/%m/%Y %H:%M")
                    self.current_date = "Date and time: %s" % dt_string
                    pdf.drawString(80, 150, self.current_date)
                    pdf.setFont("source", 14)
                    author = ""
                    username = getpass.getuser()
                    user = "%s" % username
                    print(user)
                    if user == "kumardhivake":
                        author = "Author (Department): Dhivaker Kumar  (IFAG IPC DD C)"
                    if user == "BiswasAr":
                        author = "Author (Department): Arnab Biswas (IFAG IPC DD C)"
                    if user == "Cotoroge":
                        author = "Author (Department): Maria Cotorogea (IFAG IPC DD C)"
                    pdf.drawString(80, 180, author)
                    pdf.showPage()
                    pdf.drawImage(logo, 450, 750, width=80, height=40, mask='auto')
                    page_num = pdf.getPageNumber()
                    page = "%s" % page_num
                    pdf.setFont('source', 12)
                    pdf.drawString(100 * mm, 10 * mm, page)
                    pdf.setFont("source", 14)
                    pdf.drawString(50, 750, self.product_name+"-"+self.model_level)
                    pdf.line(50, 30 * mm, 530, 30 * mm)  # bottom line
                    pdf.line(50, 745, 530, 745)
                    pdf.setFont('source', 10)
                    pdf.drawString(50, 25 * mm, "Copyright © Infineon Technologies AG 2020. All rights reserved")
                    pdf.setFont('source', 24)
                    pdf.drawImage(os.path.join(self.dir_out_plot,'model_det.png'), 80, 340, width=15.92 * cm, height=12.74 * cm)
                    pdf.drawCentredString(290, 690, "Model")
                    pdf.drawCentredString(290, 450, "Simulator settings")
                    pdf.drawImage('simulator_settings.png', 160,200, width=10.1 * cm, height=7.31 * cm)
                    pdf.showPage()
                labels = self.get_tlabels(images)
                pdf.drawImage(os.path.join(self.dir_out_plot,images), labels["xi_orientation"], labels["yi_orientation"], width=labels["iwidth"],
                              height=labels["iheight"])
                pdf.drawImage(os.path.join(self.dir_out_plot,'table_' + images), labels["xi_orientation"] - 70, labels["yi_orientation"] - 400,
                              width=labels["iwidth"] + 150, height=labels["iheight"] + 150)
                net_fig_path = os.path.join(self.dir_cur, "fig_net")
                pdf.drawImage(os.path.join(net_fig_path, 'net_' + images), labels["xn_orientation"],
                              labels["yn_orientation"],
                              width=labels["nwidth"]*cm, height=labels["nheight"]*cm)
                pdf.setFont("source", 24)
                pdf.drawCentredString(300, 680, labels["title"])
                #pdf.setFont("Courier-Bold", 20)
                #pdf.drawCentredString(300, 650, labels["subtitle"])
                pdf.drawImage(logo, 450, 750, width=80, height=40, mask='auto')
                page_num = pdf.getPageNumber()
                page = "%s" % page_num
                pdf.setFont('source', 12)
                pdf.drawString(100 * mm, 10 * mm, page)
                pdf.setFont("source", 14)
                pdf.drawString(50, 750, self.product_name+"-"+self.model_level)
                pdf.line(50, 30 * mm, 530, 30 * mm) #bottom line
                pdf.line(50, 745, 530, 745)
                pdf.setFont('source', 10)
                pdf.drawString(50, 25 * mm, "Copyright © Infineon Technologies AG 2020. All rights reserved")
                pdf.showPage()
        end_page = 'endpage.png'
        pdf.drawImage(end_page, 0, -3, width=21.15 * cm, height=30.08 * cm)

        pdf.save()
        self.merge_pdf(pdf_folder,pdf_filename)

    def merge_pdf(self,pdf_folder,pdf_filename):
        product_name  = self.product_name
        bare_die_report = open(pdf_folder, 'rb')
        data_sheet_path =self.datasheet_loc
        data_sheet = open(data_sheet_path, 'rb')

        # Read the files that you have opened
        pdf1Reader = PyPDF2.PdfFileReader(bare_die_report)
        pdf2Reader = PyPDF2.PdfFileReader(data_sheet)

        # Create a new PdfFileWriter object which represents a blank PDF document
        pdfWriter = PyPDF2.PdfFileWriter()

        # Loop through all the pagenumbers for the first document
        for pageNum in range(pdf1Reader.numPages):
            pageObj = pdf1Reader.getPage(pageNum)
            pdfWriter.addPage(pageObj)

        # Loop through all the pagenumbers for the second document
        for pageNum in range(pdf2Reader.numPages):
            pageObj = pdf2Reader.getPage(pageNum)
            pdfWriter.addPage(pageObj)

        # Now that you have copied all the pages in both the documents, write them into the a new document
        merged_pdf_name = 'Complete_report_'+pdf_filename
        merged_file = open(os.path.join(self.dir_out_plot,merged_pdf_name), 'wb')

        pdfWriter.write(merged_file)

        # Close all the files - Created as well as opened
        merged_file.close()
        bare_die_report.close()
        data_sheet.close()
        pdf_path = self.dir_out_plot+"\\"+merged_pdf_name
        p = subprocess.Popen([pdf_path],shell=True)
        # When you want the program to close
        # p.kill()
